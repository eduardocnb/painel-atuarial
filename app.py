# app.py
# Portal com 5 módulos: Home, Duração do Passivo (E4), Fluxo Atuarial (5 linhas),
# e NOVOS: PREV_GA (dashboard de fluxos), FIN_GA (dashboard de fluxos), PREV_GF (dashboard de fluxos).

import os, io, glob
from typing import Optional, Tuple, List, Dict
from openpyxl.utils import column_index_from_string
import streamlit as st
import os, zipfile, io, shutil, tempfile
import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
import altair as alt
import re
# msoffcrypto-tool é opcional (apenas se houver senha)
try:
    import msoffcrypto  # pip install msoffcrypto-tool
except Exception:
    msoffcrypto = None

# ------------------ CONFIG GERAL ------------------
# --- INÍCIO: Configuração de dados para rodar em Streamlit Cloud ---
import os, zipfile, io, shutil, tempfile
import streamlit as st

# Onde o app vai procurar dados por padrão no Streamlit Cloud / GitHub:
# Estrutura esperada do repositório (se usar dados no repo):
#   data/
#     Fluxo/2015/*.xlsx ... Fluxo/2025/*.xlsx
#     DIPR/DIPR_filtrado/RG_OUTROS_DP_ENVIO_* (YYYY).xlsx
DEFAULT_DATA_ROOT = os.path.join(os.path.dirname(__file__), "data")

# Permitir alternar a FONTE DE DADOS pela UI:
with st.sidebar:
    st.subheader("🔧 Fonte de dados")
    fonte = st.radio(
        "Escolha a origem dos arquivos",
        options=["Pasta do repositório (./data)", "Upload de ZIP (Excel)"],
        index=0,
        help="Se os arquivos não estiverem no repositório, envie um .zip com as pastas Fluxo/ e DIPR/."
    )

    if "data_root" not in st.session_state:
        st.session_state["data_root"] = DEFAULT_DATA_ROOT

    if fonte == "Upload de ZIP (Excel)":
        zip_file = st.file_uploader("Envie um .zip contendo as pastas Fluxo/ e/ou DIPR/", type=["zip"])
        if zip_file is not None:
            # descompacta em um diretório temporário do Streamlit Cloud
            tmp_root = tempfile.mkdtemp(prefix="dados_")
            with zipfile.ZipFile(io.BytesIO(zip_file.read())) as zf:
                zf.extractall(tmp_root)
            st.session_state["data_root"] = tmp_root
            st.success("ZIP carregado e extraído com sucesso ✅")
    else:
        st.session_state["data_root"] = DEFAULT_DATA_ROOT

    st.caption(f"📂 Data root atual: `{st.session_state['data_root']}`")

# Variáveis de caminho usadas pelo restante do app:
DATA_ROOT = st.session_state["data_root"]
BASE_DIR  = os.path.join(DATA_ROOT, "Fluxo")                       # antes: C:\Users\...\Fluxo
DIPR_DIR  = os.path.join(DATA_ROOT, "DIPR", "DIPR_filtrado")       # antes: C:\Users\...\DIPR\DIPR_filtrado

# Mostra se as pastas existem (útil para diagnóstico rápido)
def _status_dir(p):
    return "✅" if os.path.isdir(p) else "⚠️"
with st.sidebar:
    st.write(f"{_status_dir(BASE_DIR)} Fluxo: `{BASE_DIR}`")
    st.write(f"{_status_dir(DIPR_DIR)} DIPR (filtrado): `{DIPR_DIR}`")
    if st.button("Recarregar dados"):
        st.cache_data.clear()
        st.rerun()
# --- FIM: Configuração de dados ---



SENHA = "2020"
ANOS = list(range(2015, 2026))
PATTERNS = {
    "PREV_GA": "{ano}_FLX_CIVIL_PREV_GA_*_COM_DURATION.xlsx",
    "PREV_GF": "{ano}_FLX_CIVIL_PREV_GF_*_COM_DURATION.xlsx",
    "FIN_GA" : "{ano}_FLX_CIVIL_FIN_GA_*_COM_DURATION.xlsx",
}
MES_MAP = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4, "MAI": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12
}


# ------------------ UTIL EXCEL ------------------
def abrir_primeira_aba(path: str, senha: Optional[str]) -> Tuple[openpyxl.Workbook, str]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        return wb, wb.sheetnames[0]
    except Exception:
        if msoffcrypto is None:
            raise
        with open(path, "rb") as f:
            off = msoffcrypto.OfficeFile(f)
            off.load_key(password=senha or "")
            bio = io.BytesIO()
            off.decrypt(bio)
        wb = openpyxl.load_workbook(bio, data_only=True, read_only=False)
        return wb, wb.sheetnames[0]

def col_vals(ws, col_letter: str, r0: int, r1: int, as_float=True) -> List[Optional[float]]:
    ci = column_index_from_string(col_letter)
    out = []
    for r in range(r0, r1 + 1):
        v = ws.cell(row=r, column=ci).value
        if not as_float:
            out.append(v)
            continue
        if v is None or v == "":
            out.append(0.0)
        elif isinstance(v, (int, float)):
            out.append(float(v))
        else:
            s = str(v).strip().replace(".", "").replace(",", ".")
            try:
                out.append(float(s))
            except Exception:
                out.append(0.0)
    return out

def sum_range_row(ws, c_from: str, c_to: str, r0: int, r1: int, weight: Optional[List[float]]=None) -> List[float]:
    c1 = column_index_from_string(c_from)
    c2 = column_index_from_string(c_to)
    out = []
    for i, r in enumerate(range(r0, r1 + 1)):
        acc = 0.0
        w = 1.0 if weight is None else float(weight[i] or 0.0)
        for c in range(c1, c2 + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or v == "":
                continue
            if isinstance(v, (int, float)):
                acc += float(v) * (w if weight is not None else 1.0)
            else:
                s = str(v).strip().replace(".", "").replace(",", ".")
                try:
                    acc += float(s) * (w if weight is not None else 1.0)
                except Exception:
                    pass
        out.append(acc)
    return out

def col_times_weight(ws, col_letter: str, r0: int, r1: int, weight: Optional[List[float]]) -> List[float]:
    vals = col_vals(ws, col_letter, r0, r1)
    out = []
    for i, v in enumerate(vals):
        if weight is None:
            out.append(v)
        else:
            w = float(weight[i] or 0.0)
            out.append(float(v) * w)
    return out


def _letters_to_usecols(letters):
    # mapeia letras ('B','C','D',...) para índices 0-based do pandas (header=None)
    idxs = [column_index_from_string(L) - 1 for L in letters]
    return idxs
def _read_first_existing(*candidates):
    for p in candidates:
        if os.path.exists(p):
            return p
    return None
from openpyxl.utils import column_index_from_string

from openpyxl.utils import column_index_from_string
import pandas as pd, os, glob, re

# --- helpers ---
def _to_float_br(x):
    """Converte '1.234.567,89' -> 1234567.89; vazio -> 0.0"""
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "" or s.upper() in {"NAN", "NONE"}:
        return 0.0
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0

def _letters_to_idx(letters):
    """Converte letras de coluna Excel para índices 0-based (pandas header=None)."""
    return [column_index_from_string(L) - 1 for L in letters]

MES_MAP = {"JAN":1, "FEV":2, "MAR":3, "ABR":4, "MAI":5, "JUN":6,
           "JUL":7, "AGO":8, "SET":9, "OUT":10, "NOV":11, "DEZ":12}

def _parse_mesano(texto: str):
    if not isinstance(texto, str): 
        return None
    t = texto.strip().upper()
    m = re.match(r"([A-ZÇ]{3})/(\d{4})", t)
    if not m: 
        return None
    mes = MES_MAP.get(m.group(1))
    if not mes:
        return None
    return pd.Timestamp(year=int(m.group(2)), month=mes, day=1)

# --- LOADER DIPR (usa DIPR_DIR diretamente) ---
def ler_dipr_ano(ano: int, colunas_letras: list[str]) -> pd.DataFrame:
    """
    Retorna DF com colunas: ['data','tipo', <colunas_letras...>]
    - Lê arquivos da pasta DIPR_DIR (já deve apontar para DIPR_filtrado)
    - Apenas linhas B == 'DF'
    - 'data' em C (ex.: JAN/2024)
    - 'tipo' em D (Financeiro / Previdenciário)
    - Colunas escolhidas convertidas para float (pt-BR -> float)
    """
    # procura diretamente em DIPR_DIR
    patt = os.path.join(DIPR_DIR, f"RG_OUTROS_DP_ENVIO_* ({ano}).xls*")
    hits = glob.glob(patt)
    if not hits:
        return pd.DataFrame(columns=["data", "tipo"] + colunas_letras)
    path = sorted(hits)[0]

    # B,C,D + solicitadas (sem duplicar)
    base_cols = ["B", "C", "D"]
    letras_all = base_cols + [c for c in colunas_letras if c not in base_cols]
    usecols = _letters_to_idx(letras_all)

    # lê sem cabeçalho, como texto (para preservar vírgulas)
    df = pd.read_excel(
        path,
        sheet_name=0,
        header=None,
        dtype=str,
        na_filter=False,
        usecols=usecols,
        engine="openpyxl",
    )
    # renomeia para as letras correspondentes
    df.columns = letras_all[:len(df.columns)]

    # filtra somente DF (coluna B)
    df = df[df["B"].astype(str).str.strip().str.upper().eq("DF")]

    # data e tipo
    df["data"] = df["C"].apply(_parse_mesano)
    df = df.dropna(subset=["data"])
    df["tipo"] = (
        df["D"].astype(str).str.strip().str.lower()
          .map({"financeiro": "Financeiro",
                "previdenciário": "Previdenciário",
                "previdenciario": "Previdenciário"})
          .fillna("Financeiro")
    )

    # seleciona e converte colunas numéricas
    out = df[["data", "tipo"] + colunas_letras].copy()
    for c in colunas_letras:
        out[c] = out[c].map(_to_float_br)

    return out








def resultado_atuarial_series(ws, plano: str, incluir_rentab: bool, r0: int, r1: int) -> List[float]:
    """
    Retorna a série 'Resultado Atuarial' para a 1ª aba do arquivo aberto (ws).
    Regras:
      - FIN_GA: sem ponderação por D  → (-AQ) + BK [+ BN]
      - PREV_*: ponderado por D       → (-AQ*D) + (BK*D) [+ (BN*D)]
      - incluir_rentab=True → soma BN (ponderado nos PREV_*)
    """
    # colunas base
    AQ = col_vals(ws, "AQ", r0, r1)       # sinal será invertido
    BK = col_vals(ws, "BK", r0, r1)
    BN = col_vals(ws, "BN", r0, r1)       # rentabilidade dos ativos
    D  = col_vals(ws, "D",  r0, r1) if plano in ("PREV_GA", "PREV_GF") else None

    out = []
    for i in range(r1 - r0 + 1):
        aq = -(AQ[i] or 0.0)
        bk =  (BK[i] or 0.0)
        bn =  (BN[i] or 0.0) if incluir_rentab else 0.0

        if D is None:  # FIN_GA
            v = aq + bk + bn
        else:          # PREV_*
            w = float(D[i] or 0.0)
            v = (aq * w) + (bk * w) + (bn * w)

        out.append(v)
    return out


# ------------------ DURAÇÃO DO PASSIVO (E4) ------------------
def ler_e4_series(tipo: str) -> pd.DataFrame:
    CEL = "E4"
    registros = []
    for ano in ANOS:
        pasta = os.path.join(BASE_DIR, str(ano))
        if not os.path.isdir(pasta):
            continue
        pad = os.path.join(pasta, PATTERNS[tipo].format(ano=ano))
        files = glob.glob(pad)
        if not files:
            continue
        caminho = sorted(files)[0]
        try:
            wb, aba = abrir_primeira_aba(caminho, SENHA)
            sh = wb[aba]
            val = sh[CEL].value
            wb.close()
            if val is None:
                continue
            if isinstance(val, (int, float)):
                v = round(float(val), 2)
            else:
                s = str(val).strip().replace(".", "").replace(",", ".")
                v = round(float(s), 2)
            registros.append({"Ano": ano, "Serie": tipo, "E4": v})
        except Exception:
            pass
    return pd.DataFrame(registros)

def chart_e4(df: pd.DataFrame, titulo: str):
    domain_x = ANOS
    if df.empty:
        st.info(f"Sem dados para {titulo}.")
        return

    # adiciona rótulo de status para criar legenda
    df_plot = df.sort_values("Ano").copy()
    df_plot["Status"] = "Com dado"

    present = df_plot["Ano"].tolist()
    maxy = float(df_plot["E4"].max()) if not df_plot["E4"].empty else 1.0
    ytop = maxy * 1.03
    miss = [a for a in domain_x if a not in present]
    miss_df = pd.DataFrame({"Ano": miss, "E4": [ytop]*len(miss), "Status": ["Sem dado"]*len(miss)})

    # paleta/legenda (linha azul; X vermelho)
    color_enc = alt.Color(
        "Status:N",
        title="Legenda",
        scale=alt.Scale(domain=["Com dado", "Sem dado"], range=["#1f77b4", "red"]),
        legend=alt.Legend(symbolType="stroke")
    )

    line = (
        alt.Chart(df_plot)
        .mark_line(point=True)
        .encode(
            x=alt.X("Ano:O", title="Ano", scale=alt.Scale(domain=domain_x)),
            y=alt.Y("E4:Q", title="E4 (duração)", axis=alt.Axis(format=".2f")),
            color=color_enc,
            tooltip=[alt.Tooltip("Ano:O"), alt.Tooltip("E4:Q", format=".2f")],
        )
    )

    crosses = (
        alt.Chart(miss_df)
        .mark_point(shape="cross", size=140, filled=False)  # X vermelho
        .encode(
            x=alt.X("Ano:O", scale=alt.Scale(domain=domain_x)),
            y=alt.Y("E4:Q"),
            color=color_enc,
            tooltip=[alt.Tooltip("Ano:O", title="Ano"), alt.Tooltip("Status:N")],
        )
    )

    st.altair_chart((line + crosses).properties(title=titulo, height=300), use_container_width=True)


# ------------------ FLUXO ATUARIAL (5 LINHAS) ------------------
def ler_fluxo_atuarial_basico(path: str, tipo: str) -> pd.DataFrame:
    """5 curvas padrão (BC, Apos Atuais, Pens Atuais, Apos Futuros, Pens Futuros) para um arquivo."""
    wb, aba = abrir_primeira_aba(path, SENHA)
    ws = wb[aba]
    r0 = 10
    r1 = 159 if tipo == "PREV_GF" else 109  # 150 anos GF; 100 anos GA
    anos = col_vals(ws, "B", r0, r1, as_float=True)
    anos = [int(a) for a in anos if a is not None]
    # séries (sem ponderação aqui; é o painel antigo)
    base_calc = col_vals(ws, "E", r0, r1)
    apos_at = sum_range_row(ws, "AT", "AW", r0, r1)  # concedidos
    pens_at = col_vals(ws, "AX", r0, r1)
    apos_fu = sum_range_row(ws, "BA", "BD", r0, r1)
    pens_fu = sum_range_row(ws, "BE", "BF", r0, r1)
    wb.close()
    dados = pd.DataFrame({
        "AnoX": list(range(len(base_calc))),
        "Base Cálculo Contrib. Normal (BC)": base_calc,
        "Aposentados Atuais (BC)": apos_at,
        "Pensionistas Atuais (BC)": pens_at,
        "Aposentados Futuros (BaC)": apos_fu,
        "Pensionistas Futuros (BcF)": pens_fu,
    })
    dados["AnoX"] = col_vals(ws, "B", r0, r1, as_float=False)[:len(dados)]
    dados = dados.dropna(subset=["AnoX"])
    return dados.melt(id_vars=["AnoX"], var_name="Linha", value_name="Valor")

# ------------------ NOVO: DASHBOARDS POR PLANO ------------------
# Mapeamento dos "Fluxos Disponíveis"
FLUXOS_LISTA = [
    "Salários Futuros",
    "Benefícios Futuros (BaC+BC)",
    "Benefícios Futuros (BaC)",
    "Benefícios Futuros (BC)",
    "Contribuições Futuras (BaC+BC)",
    "Contribuições Futuras (BaC)",
    "Contribuições Futuras (BC)",
    "COMPREV Rec. (BaC+BC)",
    "COMPREV Rec. (BaC)",
    "COMPREV Rec. (BC)",
    "Provisão Matemática (BaC+BC)",
    "Provisão Matemática (BaC)",
    "Provisão Matemática (BC)",
    "Ativos Garantidores",
    "Aposentadorias (BaC+BC)",
    "Aposentadorias (BaC)",
    "Aposentadorias (BC)",
    "Pensões (BaC+BC)",
    "Pensões (BaC)",
    "Pensões (BC)",
]

# Colunas pedidas (letra -> rótulo)
DIPR_COLS = {
    "F":  '1 - BASES DE CÁLCULO... (1.1 - **Do ENTE ("patronal")**, relativa: **a) Aos Servidores**)',
    "J":  '1 - BASES DE CÁLCULO... (**1.2 - Dos Servidores**)',
    "M":  '2 - CONTRIBUIÇÕES REPASSADAS (**a) Aos Servidores**)',
    "Q":  '2 - CONTRIBUIÇÕES REPASSADAS (**2.2 - Dos Servidores**)',
    "R":  '2 - CONTRIBUIÇÕES REPASSADAS (**2.3 - Dos Aposentados**)',
    "S":  '2 - CONTRIBUIÇÕES REPASSADAS (**2.4 - Dos Pensionistas**)',
    "W":  '3 - DEDUÇÕES (3.2 - Valores deduzidos...) **b) Outros valores Compensados**',
    "Y":  '4 - APORTES... (**4.2 - Transferência p/ Insuficiência Financeira**)',
    "Z":  '4 - APORTES... (**4.3 - Transf. p/ despesas administrativas**)',
    "AA": '4 - APORTES... (**4.4 - Transf. p/ benefícios do Tesouro**)',
    "AB": '4 - APORTES... (**4.5 - Outros aportes/transferências**)',
    "AD": '6 - BASES... (**6.1 - Da UNIDADE GESTORA**)',
    "AE": '6 - BASES... (**6.2 - Dos SERVIDORES da UG**)',
    "AG": '6 - BASES... (**6.4 - Dos APOSENTADOS**)',
    "AH": '6 - BASES... (**6.5 - Dos PENSIONISTAS**)',
    "AI": '7 - CONTRIBUIÇÕES ARRECADADAS (**7.1 - Da UG**)',
    "AJ": '7 - CONTRIBUIÇÕES ARRECADADAS (**7.2 - Dos SERVIDORES da UG**)',
    "AL": '7 - CONTRIBUIÇÕES ARRECADADAS (**7.4 - Dos APOSENTADOS**)',
    "AM": '7 - CONTRIBUIÇÕES ARRECADADAS (**7.5 - Dos PENSIONISTAS**)',
    "AN": '8 - REMUNERAÇÃO BRUTA (**8.1 - Dos SERVIDORES**)',
    "AO": '8 - REMUNERAÇÃO BRUTA (**8.2 - Dos APOSENTADOS**)',
    "AP": '8 - REMUNERAÇÃO BRUTA (**8.3 - Dos PENSIONISTAS**)',
    "AQ": '9 - Nº DE BENEFICIÁRIOS (**9.1 - SERVIDORES**)',
    "AR": '9 - Nº DE BENEFICIÁRIOS (**9.2 - APOSENTADOS**)',
    "AS": '9 - Nº DE BENEFICIÁRIOS (**9.3 - PENSIONISTAS**)',
    "AT": '9 - Nº DE BENEFICIÁRIOS (**9.4 - DEPENDENTES**)',
    "AU": '10 - INGRESSOS (**10.1 - Contribuições**)',
    "AV": '10 - INGRESSOS (**10.2 - Aportes**)',
    "AY": '10 - INGRESSOS (**10.5 - Contribuições - cedidos/licenciados**)',
    "AZ": '10 - INGRESSOS (**10.6 - Compensação Financeira**)',
    "BA": '10 - INGRESSOS (**10.7 - Receitas líquidas aplicações**)',
    "BB": '10 - INGRESSOS (**10.8 - Rendimento demais ativos**)',
    "BC": '10 - INGRESSOS (**10.9 - Outras receitas**)',
    "BD": '11 - UTILIZAÇÃO (**11.1 - Aposentadoria**)',
    "BE": '11 - UTILIZAÇÃO (**11.2 - Pensão por morte**)',
    "BH": '11 - UTILIZAÇÃO (**11.5 - Salário-família**)',
    "BJ": '11 - UTILIZAÇÃO (**11.7 - Decisões judiciais (benefícios)**)',
    "BK": '11 - UTILIZAÇÃO (**11.8 - Benefícios do Tesouro**)',
    "BL": '11 - UTILIZAÇÃO (**11.9 - Despesas Administrativas**)',
    "BM": '11 - UTILIZAÇÃO (**11.10 - Investimentos**)',
    "BN": '11 - UTILIZAÇÃO (**11.11 - Restituições/compensações pagas**)',
    "BO": '11 - UTILIZAÇÃO (**11.12 - Compensação Financeira (pagamento)**)',
    "BP": '11 - UTILIZAÇÃO (**11.13 - Outras despesas**)',
    "BQ": '12 - RESULTADO FINAL (**12.1 - TOTAL DE INGRESSOS**)',
    "BR": '12 - RESULTADO FINAL (**12.2 - TOTAL DE UTILIZAÇÃO**)',
    "BS": '12 - RESULTADO FINAL (**12.3 - RESULTADO FINAL APURADO**)',
}
INF_FACTOR_2025 = {
    2015: 1.7625,
    2016: 1.5954,
    2017: 1.4912,
    2018: 1.4505,
    2019: 1.3941,
    2020: 1.3499,
    2021: 1.2941,
    2022: 1.1686,
    2023: 1.1035,
    2024: 1.0541,
    2025: 1.0000,
}
# --- Correção pela inflação (base 2025) ---
# Percentuais informados pelo Eduardo. Fator = 1 + percentual.
INF_PCT_2025 = {
    2015: 0.7625,
    2016: 0.5954,
    2017: 0.4912,
    2018: 0.4505,
    2019: 0.3941,
    2020: 0.3499,
    2021: 0.2941,
    2022: 0.1686,
    2023: 0.1035,
    2024: 0.0541,
    2025: 0.0,
}
def infl_factor_2025(ano_arquivo: int) -> float:
    return 1.0 + INF_PCT_2025.get(ano_arquivo, 0.0)
def infl_factor_from_year(ano: int) -> float:
    return float(INF_FACTOR_2025.get(int(ano), 1.0))

def dipr_anos_disponiveis():
    anos = []
    for p in glob.glob(os.path.join(DIPR_DIR, "RG_OUTROS_DP_ENVIO_* (*.xls*")):
        m = re.search(r"\((\d{4})\)", os.path.basename(p))
        if m:
            anos.append(int(m.group(1)))
    return sorted(set(anos))

def _parse_mesano(texto: str) -> Optional[pd.Timestamp]:
    if not isinstance(texto, str):
        return None
    t = texto.strip().upper()
    m = re.match(r"([A-ZÇ]{3})/(\d{4})", t)
    if not m: 
        return None
    mes = MES_MAP.get(m.group(1))
    if not mes: 
        return None
    ano = int(m.group(2))
    return pd.Timestamp(year=ano, month=mes, day=1)




def dipr_series(conjunto_anos: List[int], tipo_sel: str, cols_escolhidas: List[str], corrigir: bool) -> pd.DataFrame:
    """
    Retorna DF longo com colunas: data, valor, serie
    - concatena todos os anos (mensal)
    - filtra por tipo_sel ('Financeiro' ou 'Previdenciário' ou 'Todos')
    - se corrigir=True, duplica cada série com sufixo ' (corrigido)' multiplicando pelo fator (base 2025)
    """
    frames = []
    for ano in conjunto_anos:
        base = ler_dipr_ano(ano, cols_escolhidas)
        if base.empty: 
            continue
        if tipo_sel != "Todos":
            base = base[base["tipo"].eq(tipo_sel)]
        if base.empty:
            continue

        # longo (uma série por coluna)
        long = base.melt(id_vars=["data","tipo"], value_vars=cols_escolhidas, var_name="col", value_name="valor")
        # mapeia rótulos bonitos
        long["serie"] = long["col"].map(DIPR_COLS)
        long.drop(columns=["col"], inplace=True)

        # inflação (base 2025) — usa o ano do ARQUIVO
        if corrigir:
            adj = long.copy()
            adj["fator_inf"] = adj["data"].dt.year.map(infl_factor_from_year).fillna(1.0)
            adj["valor"] = adj["valor"] * adj["fator_inf"]
            adj["serie"] = adj["serie"] + " (corrigido)"
            adj = adj.drop(columns=["fator_inf"])
            frames.append(adj)

        frames.append(long)

    if not frames:
        return pd.DataFrame(columns=["data","valor","serie"])

    out = pd.concat(frames, ignore_index=True)
    # ordena por data
    out = out.sort_values("data")
    return out

def dipr_series_anual(conjunto_anos: List[int], tipo_sel: str, cols_escolhidas: List[str], corrigir: bool) -> pd.DataFrame:
    """Agrega por ANO (soma dos meses)."""
    mens = dipr_series(conjunto_anos, tipo_sel, cols_escolhidas, corrigir)
    if mens.empty:
        return mens
    mens["Ano"] = mens["data"].dt.year
    return (
        mens.groupby(["Ano","serie"], as_index=False)["valor"].sum()
        .sort_values(["serie","Ano"])
    )


def ler_fluxo_por_nome(ws, r0, r1, plano: str, nome_fluxo: str) -> List[float]:
    """
    Calcula uma série por nome. Ponderação por D somente em planos PREV_* (PREV_GA, PREV_GF).
    FIN_GA não pondera por D.
    """
    pondera = plano in ("PREV_GA", "PREV_GF")
    D = col_vals(ws, "D", r0, r1) if pondera else None

    def soma_intervalos(intervalos: List[Tuple[str, str]]) -> List[float]:
        acc = [0.0] * (r1 - r0 + 1)
        for a, b in intervalos:
            parc = sum_range_row(ws, a, b, r0, r1, weight=D)
            acc = [x + y for x, y in zip(acc, parc)]
        return acc

    # ---- Salários Futuros ----
    # FIN_GA: usar apenas coluna E (sem D)
    # PREV_*: multiplicar E por D linha a linha
    if nome_fluxo == "Salários Futuros":
        E = col_vals(ws, "E", r0, r1)
        if plano == "FIN_GA":
            return E
        # PREV_*:
        return [ (float(d or 0.0) * float(e or 0.0)) for d, e in zip(D or [0.0]*len(E), E) ]

    # ---- Benefícios Futuros ----
    if nome_fluxo == "Benefícios Futuros (BaC)":
        return soma_intervalos([("BA", "BI")])
    if nome_fluxo == "Benefícios Futuros (BC)":
        return sum_range_row(ws, "AT", "AY", r0, r1, weight=D)
    if nome_fluxo == "Benefícios Futuros (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Benefícios Futuros (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Benefícios Futuros (BC)")
        return [x + y for x, y in zip(a, b)]

    # ---- Contribuições Futuras ----
    if nome_fluxo == "Contribuições Futuras (BaC)":
        return soma_intervalos([("N","T"), ("V","AB"), ("AD","AG"), ("AI","AM")])
    if nome_fluxo == "Contribuições Futuras (BC)":
        return sum_range_row(ws, "G", "K", r0, r1, weight=D)
    if nome_fluxo == "Contribuições Futuras (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Contribuições Futuras (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Contribuições Futuras (BC)")
        return [x + y for x, y in zip(a, b)]

    # ---- COMPREV ----
    if nome_fluxo == "COMPREV Rec. (BaC)":
        AN = col_times_weight(ws, "AN", r0, r1, D)
        BH = col_times_weight(ws, "BH", r0, r1, D)
        return [x - y for x, y in zip(AN, BH)]
    if nome_fluxo == "COMPREV Rec. (BC)":
        L = col_times_weight(ws, "L", r0, r1, D)
        AY = col_times_weight(ws, "AY", r0, r1, D)
        return [x - y for x, y in zip(L, AY)]
    if nome_fluxo == "COMPREV Rec. (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "COMPREV Rec. (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "COMPREV Rec. (BC)")
        return [x + y for x, y in zip(a, b)]

    # ---- Provisão Matemática ----
    if nome_fluxo == "Provisão Matemática (BaC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Benefícios Futuros (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Contribuições Futuras (BaC)")
        return [x + y for x, y in zip(a, b)]
    if nome_fluxo == "Provisão Matemática (BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Benefícios Futuros (BC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Contribuições Futuras (BC)")
        return [x + y for x, y in zip(a, b)]
    if nome_fluxo == "Provisão Matemática (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Provisão Matemática (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Provisão Matemática (BC)")
        return [x + y for x, y in zip(a, b)]

    # ---- Outros ----
    if nome_fluxo == "Ativos Garantidores":
        return col_vals(ws, "BO", r0, r1)
    if nome_fluxo == "Aposentadorias (BaC)":
        return sum_range_row(ws, "BA", "BD", r0, r1, weight=D)
    if nome_fluxo == "Aposentadorias (BC)":
        return sum_range_row(ws, "AT", "AW", r0, r1, weight=D)
    if nome_fluxo == "Aposentadorias (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Aposentadorias (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Aposentadorias (BC)")
        return [x + y for x, y in zip(a, b)]
    if nome_fluxo == "Pensões (BaC)":
        return sum_range_row(ws, "BE", "BF", r0, r1, weight=D)
    if nome_fluxo == "Pensões (BC)":
        return col_times_weight(ws, "AX", r0, r1, D)
    if nome_fluxo == "Pensões (BaC+BC)":
        a = ler_fluxo_por_nome(ws, r0, r1, plano, "Pensões (BaC)")
        b = ler_fluxo_por_nome(ws, r0, r1, plano, "Pensões (BC)")
        return [x + y for x, y in zip(a, b)]

    return [0.0] * (r1 - r0 + 1)


def anos_disponiveis(plano: str) -> List[int]:
    disp = []
    for ano in ANOS:
        pasta = os.path.join(BASE_DIR, str(ano))
        if not os.path.isdir(pasta):
            continue
        if glob.glob(os.path.join(pasta, PATTERNS[plano].format(ano=ano))):
            disp.append(ano)
    return disp

def pagina_dashboard_plano(plano: str, titulo: str):
    st.header(titulo)
    st.caption("PREV_* pondera por D; FIN_GA não pondera. B define o eixo dos anos. Opcional: correção pela inflação (base 2025).")

    r1 = 159 if plano == "PREV_GF" else 109
    anos_disp = anos_disponiveis(plano)
    if not anos_disp:
        st.warning("Nenhum arquivo encontrado para este plano.")
        st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)
        return

    cA, cB, cC = st.columns([2, 2, 2])
    with cA:
        anos_sel = st.multiselect("Anos para comparação", [str(a) for a in anos_disp], default=[str(anos_disp[-1])])
    with cB:
        fluxo_sel = st.selectbox("Fluxos Disponíveis", FLUXOS_LISTA, index=1)
    with cC:
        corrigir_infl = st.checkbox("Corrigir pela inflação (base 2025)", value=False)

    if not anos_sel:
        st.info("Selecione ao menos um ano.")
    else:
        frames = []
        for a_str in anos_sel:
            ano = int(a_str)
            caminho = sorted(glob.glob(os.path.join(BASE_DIR, str(ano), PATTERNS[plano].format(ano=ano))))[0]
            wb, aba = abrir_primeira_aba(caminho, SENHA)
            ws = wb[aba]
            anos_eixo = col_vals(ws, "B", 10, r1, as_float=False)
            serie = ler_fluxo_por_nome(ws, 10, r1, plano, fluxo_sel)
            wb.close()

            # aplica correção pela inflação (base 2025) se ligado
            if corrigir_infl:
                fator = infl_factor_2025(ano)   # 1 + percentual
                serie = [v * fator for v in serie]

            df = pd.DataFrame({"AnoEixo": anos_eixo[:len(serie)], "Valor": serie, "AnoArquivo": str(ano)})
            df = df[pd.notna(df["AnoEixo"])]
            frames.append(df)

        big = pd.concat(frames, ignore_index=True)
        chart = (
            alt.Chart(big)
            .mark_line(point=True)
            .encode(
                x=alt.X("AnoEixo:O", title="Ano"),
                y=alt.Y("Valor:Q", title="Valor", axis=alt.Axis(format=",.2f")),
                color=alt.Color("AnoArquivo:N", title="Ano (arquivo)"),
                tooltip=[alt.Tooltip("AnoEixo:O", title="Ano"),
                         alt.Tooltip("AnoArquivo:N", title="Arquivo"),
                         alt.Tooltip("Valor:Q", format=",.2f")]
            )
            .properties(title=f"Comparação de Fluxos – {fluxo_sel}{' (corrigido inflação 2025)' if corrigir_infl else ''}", height=380)
        )
        st.altair_chart(chart, use_container_width=True)

        st.markdown("**Observação:** Contribuições e Benefícios considerados **sem COMPREV**; idem **Provisão Matemática**.")
        csv = big.to_csv(index=False).encode("utf-8")
        st.download_button("⬇️ Baixar CSV do gráfico", data=csv, file_name=f"{plano}_{fluxo_sel.replace(' ','_')}{'_corrigido' if corrigir_infl else ''}.csv", mime="text/csv")

    st.markdown("---")
    st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)


# ------------------ NAVEGAÇÃO ------------------
st.set_page_config(page_title="Portal – Fluxos Atuariais", layout="wide")
if "page" not in st.session_state:
    st.session_state.page = "home"

def go(p): st.session_state.page = p

if st.session_state.page == "home":
    st.title("🏛️ Portal – Indicadores Atuariais")
    st.write("Escolha um módulo.")
    c1, c2, c3 = st.columns(3)
    with c1:
        st.button("📈 Duração do Passivo", use_container_width=True, on_click=go, args=("duracao",))
        st.button("📊 Fluxo Atuarial (5 linhas)", use_container_width=True, on_click=go, args=("fluxo_basico",))
    with c2:
        st.button("🟢 Plano Previdenciário – Geração Atual", use_container_width=True, on_click=go, args=("prev_ga",))
        st.button("🔵 Plano Financeiro – Geração Atual", use_container_width=True, on_click=go, args=("fin_ga",))
    with c3:
        st.button("🟣 Plano Previdenciário – Geração Futura", use_container_width=True, on_click=go, args=("prev_gf",))
        st.button("🧮 Resultado Atuarial", use_container_width=True, on_click=go, args=("resultado_atuarial",))
        st.button("📅 DIPR mensal", use_container_width=True, on_click=go, args=("dipr_mensal",))
        st.button("📆 DIPR anual", use_container_width=True, on_click=go, args=("dipr_anual",))

elif st.session_state.page == "duracao":
    st.header("📈 Duração do Passivo (E4)")
    df_prev_ga = ler_e4_series("PREV_GA")
    df_prev_gf = ler_e4_series("PREV_GF")
    df_fin_ga  = ler_e4_series("FIN_GA")
    c1, c2, c3 = st.columns(3)
    with c1: chart_e4(df_prev_ga, "Previdenciário – GA")
    with c2: chart_e4(df_prev_gf, "Previdenciário – GF")
    with c3: chart_e4(df_fin_ga , "Financeiro – GA")
    st.markdown("---"); st.button("⬅️ Voltar", on_click=go, args=("home",), use_container_width=True)


elif st.session_state.page == "resultado_atuarial":
    st.header("🧮 Resultado Atuarial")
    st.caption("Série construída da 1ª aba. GA = 100 anos; GF = 150 anos. PREV_* pondera por D; FIN_GA não pondera. Opcional: corrigir pela inflação (base 2025) e incluir rentabilidade (BN).")

    tipos_opt = ["PREV_GA", "PREV_GF", "FIN_GA"]
    tipo_sel  = st.selectbox("Tipo", tipos_opt, index=0)

    anos_disp = anos_disponiveis(tipo_sel)
    if not anos_disp:
        st.warning("Nenhum arquivo disponível para o tipo selecionado.")
        st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)
    else:
        colA, colB, colC = st.columns([2, 2, 2])
        with colA:
            anos_sel = st.multiselect("Anos (somente os disponíveis)", [str(a) for a in anos_disp],
                                      default=[str(anos_disp[-1])])
        with colB:
            corrigir_infl = st.checkbox("Corrigir pela inflação (base 2025)", value=False)
        with colC:
            incluir_rentab = st.checkbox("Incluir Rentabilidade dos Ativos (BN)", value=True)

        if not anos_sel:
            st.info("Selecione ao menos um ano.")
        else:
            r0, r1 = 10, (159 if tipo_sel == "PREV_GF" else 109)
            frames = []
            for a_str in anos_sel:
                ano = int(a_str)
                caminho = sorted(glob.glob(os.path.join(BASE_DIR, str(ano), PATTERNS[tipo_sel].format(ano=ano))))[0]
                wb, aba = abrir_primeira_aba(caminho, SENHA)
                ws = wb[aba]

                anos_eixo = col_vals(ws, "B", r0, r1, as_float=False)
                serie = resultado_atuarial_series(ws, tipo_sel, incluir_rentab, r0, r1)
                wb.close()

                if corrigir_infl:
                    fator = infl_factor_2025(ano)
                    serie = [v * fator for v in serie]

                df = pd.DataFrame({"AnoEixo": anos_eixo[:len(serie)], "Valor": serie, "AnoArquivo": str(ano)})
                df = df[pd.notna(df["AnoEixo"])]
                frames.append(df)

            big = pd.concat(frames, ignore_index=True)
            chart = (
                alt.Chart(big)
                .mark_line(point=True)
                .encode(
                    x=alt.X("AnoEixo:O", title="Ano"),
                    y=alt.Y("Valor:Q", title="Valor", axis=alt.Axis(format=",.2f")),
                    color=alt.Color("AnoArquivo:N", title="Ano (arquivo)"),
                    tooltip=[alt.Tooltip("AnoEixo:O", title="Ano"),
                             alt.Tooltip("AnoArquivo:N", title="Arquivo"),
                             alt.Tooltip("Valor:Q", format=",.2f")]
                )
                .properties(
                    title=f"Resultado Atuarial – {tipo_sel}"
                          f"{' (com rentab.)' if incluir_rentab else ' (sem rentab.)'}"
                          f"{' – corrigido inflação 2025' if corrigir_infl else ''}",
                    height=380
                )
            )
            st.altair_chart(chart, use_container_width=True)

            csv = big.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️ Baixar CSV", data=csv,
                               file_name=f"resultado_atuarial_{tipo_sel}{'_rentab' if incluir_rentab else ''}{'_inflacao' if corrigir_infl else ''}.csv",
                               mime="text/csv")

    st.markdown("---")
    st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)

elif st.session_state.page == "dipr_mensal":
    st.header("📅 DIPR mensal")
    st.caption("Lê apenas linhas com B == 'DF'. Tipo na coluna D. Data em C (ex.: JAN/2024). Pode sobrepor várias colunas e tipos; opção de inflação sobrepõe a curva corrigida.")

    anos_disp = dipr_anos_disponiveis()
    if not anos_disp:
        st.warning("Nenhum arquivo DIPR encontrado em: " + DIPR_DIR)
        st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)
    else:
        c1, c2, c3 = st.columns([2,2,3])
        with c1:
            tipo_sel = st.selectbox("Tipo", ["Financeiro","Previdenciário","Todos"], index=0)
        with c2:
            anos_sel = st.multiselect("Anos (arquivos a usar)", [str(a) for a in anos_disp], default=[str(a) for a in anos_disp])
        with c3:
            cols_sel = st.multiselect("Colunas / Séries (pode escolher várias)", 
                                      [f"{k} - {DIPR_COLS[k]}" for k in DIPR_COLS.keys()],
                                      default=["F - " + DIPR_COLS["F"]])

        corrigir = st.checkbox("Mostrar curvas corrigidas pela inflação (base 2025)", value=False)

        if not anos_sel or not cols_sel:
            st.info("Selecione ao menos um ano e uma coluna.")
        else:
            letras = [x.split(" - ",1)[0] for x in cols_sel]
            anos_int = [int(a) for a in anos_sel]

            df = dipr_series(anos_int, tipo_sel, letras, corrigir)
            # df já veio de dipr_series(anos_int, tipo_sel, letras, corrigir)
            if df.empty:
                st.warning("Sem dados após filtros.")
            else:
                # 1) eixo temporal contínuo mês-a-mês entre min..max
                mind, maxd = df["data"].min(), df["data"].max()
                all_months = pd.date_range(mind, maxd, freq="MS")

                # 2) grade completa (todas as datas x todas as séries selecionadas)
                series = sorted(df["serie"].unique().tolist())
                grid = pd.MultiIndex.from_product([all_months, series], names=["data", "serie"]).to_frame(index=False)

                # 3) junta com os valores existentes (faltas viram NaN)
                merged = grid.merge(df[["data", "serie", "valor"]], on=["data", "serie"], how="left")

                # 4) separa válidos e faltantes (para marcar ❌ no topo de cada série)
                valid = merged.dropna(subset=["valor"]).copy()
                miss  = merged[merged["valor"].isna()].copy()

                # topo por série (para posicionar o ❌ um pouco acima)
                # topo por série (para posicionar o ❌ um pouco acima)
                if not valid.empty:
                    tops = valid.groupby("serie", as_index=False)["valor"].max().rename(columns={"valor": "top"})
                    miss = miss.merge(tops, on="serie", how="left")
                    miss["y_cross"] = miss["top"] * 1.02
                else:
                    miss["y_cross"] = 1.0  # fallback

                # cria a coluna de status para usar no tooltip
                miss["Status"] = "Sem dado"

                # 5) gráfico: linhas+marcadores para dados válidos; ❌ vermelhos para faltas
                base = (
                    alt.Chart(valid)
                    .mark_line(point=True)
                    .encode(
                        x=alt.X("data:T", title="Data (mês)"),
                        y=alt.Y("valor:Q", title="Valor", axis=alt.Axis(format=",.2f")),
                        color=alt.Color("serie:N", title="Série"),
                        tooltip=[
                            alt.Tooltip("data:T",  title="Data"),
                            alt.Tooltip("serie:N", title="Série"),
                            alt.Tooltip("valor:Q", title="Valor", format=",.2f"),
                        ],
                    )
                )

                crosses = (
                    alt.Chart(miss)
                    .mark_point(shape="cross", size=140, color="red")
                    .encode(
                        x=alt.X("data:T"),
                        y=alt.Y("y_cross:Q"),
                        tooltip=[
                            alt.Tooltip("data:T",   title="Data"),
                            alt.Tooltip("Status:N", title="Status"),
                        ],
                    )
                )

                st.altair_chart(
                    (base + crosses).properties(title="DIPR – mensal", height=380),
                    use_container_width=True,
                )

                # download único
                out = df.copy()
                out["data"] = out["data"].dt.strftime("%Y-%m")
                csv = out.to_csv(index=False).encode("utf-8")
                st.download_button("⬇️ Baixar CSV (mensal)", data=csv, file_name="dipr_mensal.csv", mime="text/csv")
    st.markdown("---")
    st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)

elif st.session_state.page == "dipr_anual":
    st.header("📆 DIPR anual")
    st.caption("Mesma lógica do mensal, mas somado por ano (soma dos meses).")

    anos_disp = dipr_anos_disponiveis()
    if not anos_disp:
        st.warning("Nenhum arquivo DIPR encontrado em: " + DIPR_DIR)
        st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)
    else:
        c1, c2, c3 = st.columns([2,2,3])
        with c1:
            tipo_sel = st.selectbox("Tipo", ["Financeiro","Previdenciário","Todos"], index=0, key="dipr_annual_tipo")
        with c2:
            anos_sel = st.multiselect("Anos (arquivos a usar)", [str(a) for a in anos_disp], 
                                      default=[str(a) for a in anos_disp], key="dipr_annual_anos")
        with c3:
            cols_sel = st.multiselect("Colunas / Séries (pode escolher várias)", 
                                      [f"{k} - {DIPR_COLS[k]}" for k in DIPR_COLS.keys()],
                                      default=["F - " + DIPR_COLS["F"]], key="dipr_annual_cols")

        corrigir = st.checkbox("Mostrar curvas corrigidas pela inflação (base 2025)", value=False, key="dipr_annual_corr")

        if not anos_sel or not cols_sel:
            st.info("Selecione ao menos um ano e uma coluna.")
        else:
            letras = [x.split(" - ",1)[0] for x in cols_sel]
            anos_int = [int(a) for a in anos_sel]

            df = dipr_series_anual(anos_int, tipo_sel, letras, corrigir)
            if df.empty:
                st.warning("Sem dados após filtros.")
            else:
                chart = (
                    alt.Chart(df)
                    .mark_line(point=True)
                    .encode(
                        x=alt.X("Ano:O", title="Ano"),
                        y=alt.Y("valor:Q", title="Valor", axis=alt.Axis(format=",.2f")),
                        color=alt.Color("serie:N", title="Série"),
                        tooltip=[alt.Tooltip("Ano:O"), alt.Tooltip("serie:N"), alt.Tooltip("valor:Q", format=",.2f")]
                    )
                    .properties(height=380, title="DIPR – anual (somatório dos meses)")
                )
                st.altair_chart(chart, use_container_width=True)

                csv = df.to_csv(index=False).encode("utf-8")
                st.download_button("⬇️ Baixar CSV (anual)", data=csv, file_name="dipr_anual.csv", mime="text/csv")

    st.markdown("---")
    st.button("⬅️ Voltar", on_click=lambda: st.session_state.update(page="home"), use_container_width=True)


elif st.session_state.page == "fluxo_basico":
    st.header("📊 Fluxo Atuarial – 5 Linhas")

    # Filtros
    colA, colB = st.columns([2, 3])
    with colA:
        tipo_opt = ["PREV_GA", "PREV_GF", "FIN_GA", "Todos"]
        tipo_sel = st.selectbox("Tipo", tipo_opt, index=3)
    # anos disponíveis em função do tipo (ou união quando 'Todos')
    if tipo_sel == "Todos":
        anos_disp = sorted(set().union(*[set(anos_disponiveis(t)) for t in ["PREV_GA", "PREV_GF", "FIN_GA"]]))
    else:
        anos_disp = anos_disponiveis(tipo_sel)

    with colB:
        anos_sel = st.multiselect(
            "Anos (apenas os disponíveis)",
            [str(a) for a in anos_disp],
            default=[str(a) for a in anos_disp]
        )

    if not anos_sel:
        st.info("Selecione ao menos um ano.")
        st.button("⬅️ Voltar", on_click=go, args=("home",), use_container_width=True)
    else:
        tipos_iter = ["PREV_GA", "PREV_GF", "FIN_GA"] if tipo_sel == "Todos" else [tipo_sel]
        for t in tipos_iter:
            anos_t = [int(a) for a in anos_sel if glob.glob(os.path.join(BASE_DIR, a, PATTERNS[t].format(ano=int(a))))]
            if not anos_t:
                continue
            st.subheader(t)
            for ano in sorted(anos_t):
                caminho = sorted(glob.glob(os.path.join(BASE_DIR, str(ano), PATTERNS[t].format(ano=ano))))[0]
                df_long = ler_fluxo_atuarial_basico(caminho, t)
                if df_long.empty:
                    continue
                ch = (
                    alt.Chart(df_long)
                    .mark_line(point=True)
                    .encode(
                        x=alt.X("AnoX:O", title="Ano"),
                        y=alt.Y("Valor:Q", title="Valor", axis=alt.Axis(format=",.2f")),
                        color=alt.Color("Linha:N", title="Componente"),
                        tooltip=["AnoX","Linha",alt.Tooltip("Valor:Q", format=",.2f")]
                    ).properties(title=f"Arquivo {ano}", height=300)
                )
                st.altair_chart(ch, use_container_width=True)

        st.markdown("---")
        st.button("⬅️ Voltar", on_click=go, args=("home",), use_container_width=True)


elif st.session_state.page == "prev_ga":
    pagina_dashboard_plano("PREV_GA", "🟢 Plano Previdenciário – Geração Atual (100 anos)")

elif st.session_state.page == "fin_ga":
    pagina_dashboard_plano("FIN_GA", "🔵 Plano Financeiro – Geração Atual (100 anos, sem ponderação por D)")

elif st.session_state.page == "prev_gf":
    pagina_dashboard_plano("PREV_GF", "🟣 Plano Previdenciário – Geração Futura (150 anos)")
